import argparse
import glob
import multiprocessing as mp
from functools import partial
from multiprocessing import Pool, cpu_count
from pathlib import Path

import openpyxl
import pandas as pd


def xlsx_replace(filename: Path | str, old: str, new: str) -> None:
    """Look for a string in an Excel file and replace it with another string.

    Parameters
    ----------
    filename : Path | str
        Path to the Excel file.
    old : str
        String to find.
    new : str
        String to replace with.
    """
    print(filename)
    wb = openpyxl.load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        i = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                s = ws.cell(r, c).value
                if s != None and isinstance(s, str):
                    if old in s:
                        ws.cell(r, c).value = s.replace(old, new)
                    i += 1
    wb.save(filename)


def xlsx_replace_single_multi(
    filein: Path, fileout: Path, old_new: dict[str, str]
) -> None:
    """Look for a string in an Excel file and replace it with another string.

    Parameters
    ----------
    filein : Path
        Path to the Excel file.
    fileout : Path
        Path to save the modified Excel file.
    old_new : dict[str, str]
        Dictionary with old strings as keys and new strings as values.
    """

    def process_sheet(sheet_name, wb, old_new):
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                s = ws.cell(r, c).value
                if s is not None and isinstance(s, str):
                    for old, new in old_new.items():
                        if old in s:
                            ws.cell(r, c).value = s.replace(old, new)

    wb = openpyxl.load_workbook(filein)

    # Create pool and process sheets in parallel
    with mp.Pool() as pool:
        pool.map(partial(process_sheet, wb=wb, old_new=old_new), wb.sheetnames)

    wb.save(fileout)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("old", help="Old string to replace")
    parser.add_argument("new", help="New string to replace with")
    parser.add_argument("files", nargs="+", help="Files to process")
    args = parser.parse_args()

    with Pool(processes=cpu_count()) as pool:
        all_analysis = pool.starmap(
            xlsx_replace, [(file, args.old, args.new) for file in args.files]
        )
    # for file in args.files:
    #     xlsx_replace(file, args.old, args.new)
    #     print(f"Replaced {args.old} with {args.new} in {file}")
